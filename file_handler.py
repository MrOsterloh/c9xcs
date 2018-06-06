"""

Bastian
created  23.05.2018

"""

from os import scandir
import re

import openpyxl


class File_handler():
    def __init__(self, parent):
        self.parent = parent

    def start_walk(self):
        for path in self.file_generator(self.parent.path):
            index = self.filetype_matcher(path)

            if index is not None:
                for _index in self.parent.file_types_index[index]:
                    if self.keyword_checker(path, _index):
                        self.read_data(path, _index)

    def file_generator(self, path):
        for entry in scandir(path):
            if entry.is_dir(follow_symlinks=False):
                yield from self.file_generator(entry.path)
            else:
                yield entry.path

    def filetype_matcher(self, path):
        for index, type in enumerate(self.parent.file_types):
            if re.search(type + '$', path):
                return index
        return None

    def keyword_checker(self, path, index):
         for key in self.parent.keywords[index]:
            if key in path.lower():
                return True
         return False

    def read_data(self, path, index):
        print(path)

    # def bak_excel_handler(self, path):
    #     sheetnames = ['Settings', 'Report BEV']
    #     cells = [['c69', 'e38', 'e40', 'e9', 'e8'], \
    #              ['c103', 'h259', 'h260', 'n261']]
    #     data = [['kommentar', 'name', 'Prüfstand', 'Berichtstyp', 'Zyklus'], \
    #             ['Gesamtverbrauch', 'Energie UBE', 'Energie Genutzt', 'Reichweite']]
    #
    #     result = dict()
    #     workbook = xlrd.open_workbook(path)
    #
    #     try:
    #         for sheet, cells, data in zip(sheetnames, cells, data):
    #             sheet = workbook.sheet_by_name(sheet)
    #             result += {dat: sheet.cell(cell).value for dat, cell in zip(data, cells)}
    #         return result
    #     except Exception as e:
    #         print(e)
    #         return


    def excel_handler(self, path):
        sheetnames = ['Settings', 'Report BEV']
        cells = [['c69', 'e38', 'e40', 'e9', 'e8'], \
                 ['c103', 'h259', 'h260', 'n261']]
        data = [['kommentar', 'name', 'Prüfstand', 'Berichtstyp', 'Zyklus'], \
                ['Gesamtverbrauch', 'Energie UBE', 'Energie Genutzt', 'Reichweite']]

        result = dict()
        workbook = openpyxl.load_workbook(path, read_only=True)

        try:
            for sheet, cells, data in zip(sheetnames, cells, data):
                sheet = workbook.get_sheet_by_name(sheet)
                result.update( {dat: sheet[cell].value for dat, cell in zip(data, cells)})
            print(result)
            return result
        except Exception as e:
            print(e)
            return

if __name__ == '__main__':
    import os
    path = os.getcwd()
    file_handler = File_handler(path)

    for entry in file_handler.file_generator(path):
        for file in file_handler.filetype_matcher(entry):
            file_handler.excel_handler(file_handler.keyword_matcher(file))
